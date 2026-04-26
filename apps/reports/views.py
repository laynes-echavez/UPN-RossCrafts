from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse, HttpResponse
from django.db.models import Sum, Count, Q, F
from django.db.models.functions import TruncDate
from django.utils import timezone
from datetime import timedelta, datetime
from decimal import Decimal
from io import BytesIO

from apps.authentication.decorators import role_required
from apps.sales.models import Sale, SaleItem
from apps.stock.models import Product, StockMovement
from apps.customers.models import Customer
from apps.ecommerce.models import Order

# Para exportar
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_RIGHT
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


@login_required
def dashboard(request):
    """Dashboard principal con KPIs y gráficos"""
    # Verificar que sea un usuario staff, no un cliente
    if not hasattr(request.user, 'role'):
        from django.shortcuts import redirect
        return redirect('ecommerce:catalog')
    today = timezone.now().date()
    yesterday = today - timedelta(days=1)
    
    # KPI 1: Total ventas hoy
    ventas_hoy = Sale.objects.filter(
        created_at__date=today,
        status='completed'
    ).aggregate(total=Sum('total'))['total'] or Decimal('0.00')
    
    ventas_ayer = Sale.objects.filter(
        created_at__date=yesterday,
        status='completed'
    ).aggregate(total=Sum('total'))['total'] or Decimal('0.00')
    
    if ventas_ayer > 0:
        variacion_ventas = ((ventas_hoy - ventas_ayer) / ventas_ayer) * 100
    else:
        variacion_ventas = 100 if ventas_hoy > 0 else 0
    
    # KPI 2: N° de transacciones hoy
    transacciones_hoy = Sale.objects.filter(
        created_at__date=today,
        status='completed'
    ).count()
    
    transacciones_ayer = Sale.objects.filter(
        created_at__date=yesterday,
        status='completed'
    ).count()
    
    if transacciones_ayer > 0:
        variacion_transacciones = ((transacciones_hoy - transacciones_ayer) / transacciones_ayer) * 100
    else:
        variacion_transacciones = 100 if transacciones_hoy > 0 else 0
    
    # KPI 3: Productos con stock bajo
    productos_stock_bajo = Product.objects.filter(
        is_active=True,
        stock_quantity__lte=F('min_stock_quantity')
    ).count()
    
    # KPI 4: Pedidos online pendientes
    pedidos_pendientes = Order.objects.filter(
        status__in=['pending', 'paid']
    ).count()
    
    # Últimas 10 ventas
    ultimas_ventas = Sale.objects.select_related('customer', 'user').prefetch_related('items').order_by('-created_at')[:10]
    
    # Productos con stock bajo (para alertas)
    productos_criticos = Product.objects.filter(
        is_active=True,
        stock_quantity__lte=F('min_stock_quantity')
    ).select_related('category').order_by('stock_quantity')[:10]
    
    context = {
        'ventas_hoy': ventas_hoy,
        'variacion_ventas': variacion_ventas,
        'transacciones_hoy': transacciones_hoy,
        'variacion_transacciones': variacion_transacciones,
        'productos_stock_bajo': productos_stock_bajo,
        'pedidos_pendientes': pedidos_pendientes,
        'ultimas_ventas': ultimas_ventas,
        'productos_criticos': productos_criticos,
    }
    
    return render(request, 'reports/dashboard.html', context)


@login_required
def ventas_semana_api(request):
    """API: Datos de ventas de los últimos 7 días para gráfico"""
    today = timezone.now().date()
    week_ago = today - timedelta(days=6)
    
    # Agrupar ventas por día
    ventas_por_dia = Sale.objects.filter(
        created_at__date__gte=week_ago,
        created_at__date__lte=today,
        status='completed'
    ).annotate(
        day=TruncDate('created_at')
    ).values('day').annotate(
        total=Sum('total')
    ).order_by('day')
    
    # Crear diccionario con todos los días
    dias_dict = {}
    for i in range(7):
        dia = week_ago + timedelta(days=i)
        dias_dict[dia] = 0
    
    # Llenar con datos reales
    for venta in ventas_por_dia:
        dias_dict[venta['day']] = float(venta['total'])
    
    # Preparar respuesta
    labels = []
    data = []
    dias_semana = ['Lun', 'Mar', 'Mié', 'Jue', 'Vie', 'Sáb', 'Dom']
    
    for dia in sorted(dias_dict.keys()):
        labels.append(dias_semana[dia.weekday()])
        data.append(dias_dict[dia])
    
    return JsonResponse({
        'labels': labels,
        'data': data
    })


@login_required
def top_productos_api(request):
    """API: Top 5 productos más vendidos"""
    # Obtener productos más vendidos del último mes
    mes_atras = timezone.now() - timedelta(days=30)
    
    top_productos = SaleItem.objects.filter(
        sale__created_at__gte=mes_atras,
        sale__status='completed'
    ).values(
        'product__name'
    ).annotate(
        cantidad=Sum('quantity')
    ).order_by('-cantidad')[:5]
    
    labels = [item['product__name'] for item in top_productos]
    data = [int(item['cantidad']) for item in top_productos]
    
    return JsonResponse({
        'labels': labels,
        'data': data
    })


@login_required
@role_required(['gerente', 'administrador'])
def reporte_ventas(request):
    """Reporte de ventas con filtros"""
    # Obtener filtros
    fecha_desde = request.GET.get('fecha_desde', '')
    fecha_hasta = request.GET.get('fecha_hasta', '')
    tipo = request.GET.get('tipo', 'todas')
    metodo_pago = request.GET.get('metodo_pago', 'todos')
    estado = request.GET.get('estado', 'todos')
    
    # Query base
    ventas = Sale.objects.select_related('customer', 'user').order_by('-created_at')
    
    # Aplicar filtros
    if fecha_desde:
        ventas = ventas.filter(created_at__date__gte=fecha_desde)
    
    if fecha_hasta:
        ventas = ventas.filter(created_at__date__lte=fecha_hasta)
    
    if tipo != 'todas':
        if tipo == 'presencial':
            ventas = ventas.exclude(payment_method='online')
        elif tipo == 'online':
            ventas = ventas.filter(payment_method='online')
    
    if metodo_pago != 'todos':
        ventas = ventas.filter(payment_method=metodo_pago)
    
    if estado != 'todos':
        ventas = ventas.filter(status=estado)
    
    # Calcular resumen
    resumen = ventas.aggregate(
        total_ventas=Sum('total'),
        total_descuentos=Sum('discount'),
        num_transacciones=Count('id')
    )
    
    total_ventas = resumen['total_ventas'] or Decimal('0.00')
    total_descuentos = resumen['total_descuentos'] or Decimal('0.00')
    num_transacciones = resumen['num_transacciones'] or 0
    promedio_venta = total_ventas / num_transacciones if num_transacciones > 0 else Decimal('0.00')
    
    context = {
        'ventas': ventas,
        'fecha_desde': fecha_desde,
        'fecha_hasta': fecha_hasta,
        'tipo': tipo,
        'metodo_pago': metodo_pago,
        'estado': estado,
        'total_ventas': total_ventas,
        'total_descuentos': total_descuentos,
        'num_transacciones': num_transacciones,
        'promedio_venta': promedio_venta,
    }
    
    return render(request, 'reports/ventas.html', context)


@login_required
@role_required(['gerente', 'administrador'])
def reporte_ventas_pdf(request):
    """Exportar reporte de ventas a PDF"""
    # Obtener filtros (mismos que reporte_ventas)
    fecha_desde = request.GET.get('fecha_desde', '')
    fecha_hasta = request.GET.get('fecha_hasta', '')
    tipo = request.GET.get('tipo', 'todas')
    metodo_pago = request.GET.get('metodo_pago', 'todos')
    estado = request.GET.get('estado', 'todos')
    
    # Query con filtros
    ventas = Sale.objects.select_related('customer', 'user').order_by('-created_at')
    
    if fecha_desde:
        ventas = ventas.filter(created_at__date__gte=fecha_desde)
    if fecha_hasta:
        ventas = ventas.filter(created_at__date__lte=fecha_hasta)
    if tipo == 'presencial':
        ventas = ventas.exclude(payment_method='online')
    elif tipo == 'online':
        ventas = ventas.filter(payment_method='online')
    if metodo_pago != 'todos':
        ventas = ventas.filter(payment_method=metodo_pago)
    if estado != 'todos':
        ventas = ventas.filter(status=estado)
    
    # Calcular resumen
    resumen = ventas.aggregate(
        total_ventas=Sum('total'),
        total_descuentos=Sum('discount'),
        num_transacciones=Count('id')
    )
    
    # Crear PDF
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4))
    elements = []
    styles = getSampleStyleSheet()
    
    # Estilo personalizado
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#41431B'),
        alignment=TA_CENTER,
        spaceAfter=30,
    )
    
    # Título
    elements.append(Paragraph('REPORTE DE VENTAS - ROSS CRAFTS', title_style))
    
    # Período
    periodo = f"Período: {fecha_desde or 'Inicio'} - {fecha_hasta or 'Hoy'}"
    elements.append(Paragraph(periodo, styles['Normal']))
    elements.append(Spacer(1, 20))
    
    # Tabla de ventas
    data = [['N° Venta', 'Fecha', 'Cliente', 'Método', 'Subtotal', 'Desc.', 'Total', 'Estado']]
    
    for venta in ventas[:100]:  # Limitar a 100 para no sobrecargar
        data.append([
            str(venta.id),
            venta.created_at.strftime('%d/%m/%Y'),
            venta.customer.full_name if venta.customer else 'N/A',
            venta.get_payment_method_display(),
            f"S/. {venta.subtotal:.2f}",
            f"S/. {venta.discount:.2f}",
            f"S/. {venta.total:.2f}",
            venta.get_status_display(),
        ])
    
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#41431B')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
    ]))
    
    elements.append(table)
    elements.append(Spacer(1, 20))
    
    # Resumen
    resumen_text = f"""
    <b>RESUMEN DEL PERÍODO</b><br/>
    Total ventas: S/. {resumen['total_ventas'] or 0:.2f}<br/>
    Total descuentos: S/. {resumen['total_descuentos'] or 0:.2f}<br/>
    N° transacciones: {resumen['num_transacciones'] or 0}<br/>
    Promedio por venta: S/. {(resumen['total_ventas'] or 0) / (resumen['num_transacciones'] or 1):.2f}
    """
    elements.append(Paragraph(resumen_text, styles['Normal']))
    
    # Footer
    elements.append(Spacer(1, 20))
    footer_text = f"Generado el {timezone.now().strftime('%d/%m/%Y %H:%M')} - Sistema Ross Crafts"
    elements.append(Paragraph(footer_text, styles['Normal']))
    
    # Construir PDF
    doc.build(elements)
    
    # Preparar respuesta
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    filename = f"reporte_ventas_{timezone.now().strftime('%Y%m%d')}.pdf"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response


@login_required
@role_required(['gerente', 'administrador'])
def reporte_ventas_excel(request):
    """Exportar reporte de ventas a Excel"""
    # Obtener filtros
    fecha_desde = request.GET.get('fecha_desde', '')
    fecha_hasta = request.GET.get('fecha_hasta', '')
    tipo = request.GET.get('tipo', 'todas')
    metodo_pago = request.GET.get('metodo_pago', 'todos')
    estado = request.GET.get('estado', 'todos')
    
    # Query con filtros
    ventas = Sale.objects.select_related('customer', 'user').order_by('-created_at')
    
    if fecha_desde:
        ventas = ventas.filter(created_at__date__gte=fecha_desde)
    if fecha_hasta:
        ventas = ventas.filter(created_at__date__lte=fecha_hasta)
    if tipo == 'presencial':
        ventas = ventas.exclude(payment_method='online')
    elif tipo == 'online':
        ventas = ventas.filter(payment_method='online')
    if metodo_pago != 'todos':
        ventas = ventas.filter(payment_method=metodo_pago)
    if estado != 'todos':
        ventas = ventas.filter(status=estado)
    
    # Calcular resumen
    resumen = ventas.aggregate(
        total_ventas=Sum('total'),
        total_descuentos=Sum('discount'),
        num_transacciones=Count('id')
    )
    
    # Crear workbook
    wb = Workbook()
    
    # Hoja 1: Ventas
    ws1 = wb.active
    ws1.title = "Ventas"
    
    # Estilos
    header_fill = PatternFill(start_color='41431B', end_color='41431B', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    alt_fill = PatternFill(start_color='F8F3E1', end_color='F8F3E1', fill_type='solid')
    
    # Headers
    headers = ['N° Venta', 'Fecha', 'Cliente', 'Método Pago', 'Tipo', 'Subtotal', 'Descuento', 'Total', 'Estado']
    ws1.append(headers)
    
    for cell in ws1[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # Datos
    for idx, venta in enumerate(ventas, start=2):
        ws1.append([
            venta.id,
            venta.created_at.strftime('%d/%m/%Y %H:%M'),
            venta.customer.full_name if venta.customer else 'N/A',
            venta.get_payment_method_display(),
            'Online' if venta.payment_method == 'online' else 'Presencial',
            float(venta.subtotal),
            float(venta.discount),
            float(venta.total),
            venta.get_status_display(),
        ])
        
        # Filas alternadas
        if idx % 2 == 0:
            for cell in ws1[idx]:
                cell.fill = alt_fill
    
    # Ajustar ancho de columnas
    for column in ws1.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws1.column_dimensions[column_letter].width = adjusted_width
    
    # Hoja 2: Resumen
    ws2 = wb.create_sheet("Resumen")
    ws2.append(['RESUMEN DEL PERÍODO'])
    ws2.append([])
    ws2.append(['Total ventas:', f"S/. {resumen['total_ventas'] or 0:.2f}"])
    ws2.append(['Total descuentos:', f"S/. {resumen['total_descuentos'] or 0:.2f}"])
    ws2.append(['N° transacciones:', resumen['num_transacciones'] or 0])
    ws2.append(['Promedio por venta:', f"S/. {(resumen['total_ventas'] or 0) / (resumen['num_transacciones'] or 1):.2f}"])
    
    # Estilo para resumen
    ws2['A1'].font = Font(bold=True, size=14)
    for row in ws2.iter_rows(min_row=3, max_row=6, min_col=1, max_col=1):
        for cell in row:
            cell.font = Font(bold=True)
    
    # Guardar en buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    # Preparar respuesta
    response = HttpResponse(
        buffer,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"ventas_ross_crafts_{timezone.now().strftime('%Y%m%d')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response


@login_required
@role_required(['gerente', 'administrador'])
def reporte_stock(request):
    """Reporte de stock de productos"""
    # Filtros
    categoria = request.GET.get('categoria', '')
    estado_stock = request.GET.get('estado_stock', 'todos')
    
    # Query base
    productos = Product.objects.select_related('category').filter(is_active=True)
    
    # Aplicar filtros
    if categoria:
        productos = productos.filter(category_id=categoria)
    
    if estado_stock == 'critico':
        productos = productos.filter(stock_quantity__lte=F('min_stock_quantity'))
    elif estado_stock == 'sin_stock':
        productos = productos.filter(stock_quantity=0)
    elif estado_stock == 'ok':
        productos = productos.filter(stock_quantity__gt=F('min_stock_quantity'))
    
    productos = productos.order_by('stock_quantity')
    
    # Estadísticas
    total_productos = productos.count()
    stock_critico = productos.filter(stock_quantity__lte=F('min_stock_quantity')).count()
    
    # Categorías para filtro
    from apps.stock.models import Category
    categorias = Category.objects.filter(is_active=True).order_by('name')
    
    context = {
        'productos': productos,
        'categorias': categorias,
        'categoria': categoria,
        'estado_stock': estado_stock,
        'total_productos': total_productos,
        'stock_critico': stock_critico,
    }
    
    return render(request, 'reports/stock.html', context)


@login_required
@role_required(['gerente', 'administrador'])
def reporte_stock_excel(request):
    """Exportar reporte de stock a Excel"""
    # Filtros
    categoria = request.GET.get('categoria', '')
    estado_stock = request.GET.get('estado_stock', 'todos')
    
    # Query
    productos = Product.objects.select_related('category').filter(is_active=True)
    
    if categoria:
        productos = productos.filter(category_id=categoria)
    
    if estado_stock == 'critico':
        productos = productos.filter(stock_quantity__lte=F('min_stock_quantity'))
    elif estado_stock == 'sin_stock':
        productos = productos.filter(stock_quantity=0)
    elif estado_stock == 'ok':
        productos = productos.filter(stock_quantity__gt=F('min_stock_quantity'))
    
    productos = productos.order_by('stock_quantity')
    
    # Crear workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Stock"
    
    # Estilos
    header_fill = PatternFill(start_color='41431B', end_color='41431B', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    alt_fill = PatternFill(start_color='F8F3E1', end_color='F8F3E1', fill_type='solid')
    red_fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
    
    # Headers
    headers = ['Producto', 'SKU', 'Categoría', 'Stock Actual', 'Stock Mínimo', 'Estado', 'Precio']
    ws.append(headers)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # Datos
    for idx, producto in enumerate(productos, start=2):
        estado = 'CRÍTICO' if producto.stock_quantity <= producto.min_stock_quantity else 'OK'
        
        ws.append([
            producto.name,
            producto.sku,
            producto.category.name if producto.category else 'N/A',
            producto.stock_quantity,
            producto.min_stock_quantity,
            estado,
            float(producto.price),
        ])
        
        # Resaltar críticos en rojo
        if producto.stock_quantity <= producto.min_stock_quantity:
            for cell in ws[idx]:
                cell.fill = red_fill
        elif idx % 2 == 0:
            for cell in ws[idx]:
                cell.fill = alt_fill
    
    # Ajustar ancho
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Guardar
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    response = HttpResponse(
        buffer,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"stock_ross_crafts_{timezone.now().strftime('%Y%m%d')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response


@login_required
@role_required(['gerente', 'administrador'])
def reporte_clientes(request):
    """Reporte de clientes"""
    # Top 10 clientes por monto comprado
    top_clientes = Customer.objects.annotate(
        total_comprado=Sum('sales__total', filter=Q(sales__status='completed')),
        num_compras=Count('sales', filter=Q(sales__status='completed'))
    ).filter(
        total_comprado__isnull=False
    ).order_by('-total_comprado')[:10]
    
    # Clientes nuevos por mes (últimos 6 meses)
    seis_meses_atras = timezone.now() - timedelta(days=180)
    
    clientes_por_mes = Customer.objects.filter(
        created_at__gte=seis_meses_atras
    ).annotate(
        mes=TruncDate('created_at')
    ).values('mes').annotate(
        count=Count('id')
    ).order_by('mes')
    
    # Preparar datos para gráfico
    meses_labels = []
    meses_data = []
    
    for item in clientes_por_mes:
        meses_labels.append(item['mes'].strftime('%b %Y'))
        meses_data.append(item['count'])
    
    context = {
        'top_clientes': top_clientes,
        'meses_labels': meses_labels,
        'meses_data': meses_data,
    }
    
    return render(request, 'reports/clientes.html', context)
