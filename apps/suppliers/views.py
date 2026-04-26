from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.views.generic import ListView, CreateView, UpdateView, DetailView
from django.urls import reverse_lazy
from django.http import JsonResponse, HttpResponse
from django.db.models import Q, Sum, Count
from django.db import transaction
from apps.authentication.mixins import RoleRequiredMixin
from apps.authentication.decorators import role_required
from .models import Supplier, PurchaseOrder, PurchaseOrderItem
from .forms import SupplierForm, PurchaseOrderForm
from apps.stock.models import Product, StockMovement
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
from decimal import Decimal
import json


# ==================== SUPPLIER VIEWS ====================

class SupplierListView(RoleRequiredMixin, ListView):
    """Lista de proveedores con filtros y paginación"""
    allowed_roles = ['gerente', 'administrador']
    model = Supplier
    template_name = 'suppliers/supplier_list.html'
    context_object_name = 'suppliers'
    paginate_by = 25
    
    def get_queryset(self):
        queryset = Supplier.objects.order_by('company_name')
        
        # Filtro por búsqueda
        search = self.request.GET.get('search', '')
        if search:
            queryset = queryset.filter(
                Q(company_name__icontains=search) |
                Q(contact_name__icontains=search) |
                Q(ruc__icontains=search)
            )
        
        # Filtro por estado
        is_active = self.request.GET.get('is_active', '')
        if is_active == '1':
            queryset = queryset.filter(is_active=True)
        elif is_active == '0':
            queryset = queryset.filter(is_active=False)
        
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['search'] = self.request.GET.get('search', '')
        context['selected_is_active'] = self.request.GET.get('is_active', '')
        return context


class SupplierCreateView(RoleRequiredMixin, CreateView):
    """Crear nuevo proveedor"""
    allowed_roles = ['administrador']
    model = Supplier
    form_class = SupplierForm
    template_name = 'suppliers/supplier_form.html'
    success_url = reverse_lazy('suppliers:supplier_list')
    
    def form_valid(self, form):
        messages.success(self.request, f'Proveedor "{form.instance.company_name}" creado exitosamente.')
        return super().form_valid(form)


class SupplierUpdateView(RoleRequiredMixin, UpdateView):
    """Editar proveedor existente"""
    allowed_roles = ['administrador']
    model = Supplier
    form_class = SupplierForm
    template_name = 'suppliers/supplier_form.html'
    success_url = reverse_lazy('suppliers:supplier_list')
    
    def form_valid(self, form):
        messages.success(self.request, f'Proveedor "{form.instance.company_name}" actualizado exitosamente.')
        return super().form_valid(form)


class SupplierDetailView(RoleRequiredMixin, DetailView):
    """Detalle de proveedor con historial de órdenes de compra"""
    allowed_roles = ['gerente', 'administrador']
    model = Supplier
    template_name = 'suppliers/supplier_detail.html'
    context_object_name = 'supplier'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Últimas 10 órdenes de compra
        context['purchase_orders'] = PurchaseOrder.objects.filter(
            supplier=self.object
        ).order_by('-created_at')[:10]
        
        # Estadísticas
        orders = PurchaseOrder.objects.filter(supplier=self.object)
        context['total_orders'] = orders.count()
        context['total_spent'] = orders.filter(status='received').aggregate(
            total=Sum('total')
        )['total'] or 0
        context['pending_orders'] = orders.filter(status='pending').count()
        
        return context


# ==================== PURCHASE ORDER VIEWS ====================

class PurchaseOrderListView(RoleRequiredMixin, ListView):
    """Lista de órdenes de compra con filtros"""
    allowed_roles = ['gerente', 'administrador']
    model = PurchaseOrder
    template_name = 'suppliers/purchase_order_list.html'
    context_object_name = 'orders'
    paginate_by = 25
    
    def get_queryset(self):
        queryset = PurchaseOrder.objects.select_related('supplier').order_by('-created_at')
        
        # Filtro por proveedor
        supplier_id = self.request.GET.get('supplier', '')
        if supplier_id:
            queryset = queryset.filter(supplier_id=supplier_id)
        
        # Filtro por estado
        status = self.request.GET.get('status', '')
        if status:
            queryset = queryset.filter(status=status)
        
        # Filtro por rango de fechas
        date_from = self.request.GET.get('date_from', '')
        date_to = self.request.GET.get('date_to', '')
        if date_from:
            queryset = queryset.filter(created_at__gte=date_from)
        if date_to:
            queryset = queryset.filter(created_at__lte=date_to)
        
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['suppliers'] = Supplier.objects.filter(is_active=True).order_by('company_name')
        context['selected_supplier'] = self.request.GET.get('supplier', '')
        context['selected_status'] = self.request.GET.get('status', '')
        context['date_from'] = self.request.GET.get('date_from', '')
        context['date_to'] = self.request.GET.get('date_to', '')
        return context


class PurchaseOrderCreateView(RoleRequiredMixin, CreateView):
    """Crear nueva orden de compra"""
    allowed_roles = ['administrador']
    model = PurchaseOrder
    form_class = PurchaseOrderForm
    template_name = 'suppliers/purchase_order_form.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['products'] = Product.objects.filter(is_active=True).order_by('name')
        return context
    
    def form_valid(self, form):
        try:
            with transaction.atomic():
                # Crear la orden
                order = form.save(commit=False)
                order.total = 0
                order.save()
                
                # Procesar items desde el POST
                items_data = json.loads(self.request.POST.get('items_json', '[]'))
                
                if not items_data:
                    messages.error(self.request, 'Debe agregar al menos un producto a la orden.')
                    order.delete()
                    return self.form_invalid(form)
                
                total = Decimal('0')
                
                for item_data in items_data:
                    product = Product.objects.get(id=item_data['product_id'])
                    quantity = int(item_data['quantity'])
                    unit_cost = Decimal(str(item_data['unit_cost']))
                    subtotal = quantity * unit_cost
                    
                    PurchaseOrderItem.objects.create(
                        order=order,
                        product=product,
                        quantity=quantity,
                        unit_cost=unit_cost,
                        subtotal=subtotal
                    )
                    
                    total += subtotal
                
                # Actualizar total de la orden
                order.total = total
                order.save()
                
                messages.success(self.request, f'Orden de compra OC-{order.id} creada exitosamente.')
                return redirect('suppliers:purchase_order_detail', pk=order.id)
        
        except Exception as e:
            messages.error(self.request, f'Error al crear la orden: {str(e)}')
            return self.form_invalid(form)


class PurchaseOrderDetailView(RoleRequiredMixin, DetailView):
    """Detalle de orden de compra con línea de tiempo"""
    allowed_roles = ['gerente', 'administrador']
    model = PurchaseOrder
    template_name = 'suppliers/purchase_order_detail.html'
    context_object_name = 'order'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['items'] = self.object.items.select_related('product').all()
        
        # Línea de tiempo
        timeline = []
        
        # Creación
        timeline.append({
            'status': 'created',
            'label': 'Creada',
            'date': self.object.created_at,
            'icon': '📝',
            'color': 'primary'
        })
        
        # Estado actual
        if self.object.status == 'received':
            timeline.append({
                'status': 'received',
                'label': 'Recibida',
                'date': self.object.updated_at,
                'icon': '✅',
                'color': 'success'
            })
        elif self.object.status == 'cancelled':
            timeline.append({
                'status': 'cancelled',
                'label': 'Cancelada',
                'date': self.object.updated_at,
                'icon': '❌',
                'color': 'danger'
            })
        
        context['timeline'] = timeline
        return context


@login_required
@role_required(['gerente', 'administrador'])
@transaction.atomic
def purchase_order_receive(request, pk):
    """Marcar orden de compra como recibida y actualizar stock"""
    order = get_object_or_404(PurchaseOrder, pk=pk)
    
    if order.status != 'pending':
        messages.error(request, 'Solo se pueden recibir órdenes pendientes.')
        return redirect('suppliers:purchase_order_detail', pk=pk)
    
    if request.method == 'POST':
        try:
            # Cambiar estado
            order.status = 'received'
            order.save()
            
            # Actualizar stock de cada producto
            products_updated = []
            
            for item in order.items.all():
                # Crear movimiento de stock
                StockMovement.objects.create(
                    product=item.product,
                    user=request.user,
                    movement_type='entrada',
                    quantity=item.quantity,
                    reason=f'OC #{order.id} recibida'
                )
                
                # Refrescar el objeto desde la BD para obtener el stock actualizado por la señal
                item.product.refresh_from_db()
                products_updated.append({
                    'name': item.product.name,
                    'quantity': item.quantity,
                    'new_stock': item.product.stock_quantity
                })
            
            messages.success(
                request,
                f'Orden OC-{order.id} marcada como recibida. '
                f'{len(products_updated)} productos actualizados en stock.'
            )
            
            return redirect('suppliers:purchase_order_detail', pk=pk)
        
        except Exception as e:
            messages.error(request, f'Error al recibir la orden: {str(e)}')
            return redirect('suppliers:purchase_order_detail', pk=pk)
    
    return render(request, 'suppliers/purchase_order_confirm_receive.html', {
        'order': order
    })


@login_required
@role_required(['gerente'])
def purchase_order_cancel(request, pk):
    """Cancelar orden de compra"""
    order = get_object_or_404(PurchaseOrder, pk=pk)
    
    if order.status != 'pending':
        messages.error(request, 'Solo se pueden cancelar órdenes pendientes.')
        return redirect('suppliers:purchase_order_detail', pk=pk)
    
    if request.method == 'POST':
        order.status = 'cancelled'
        order.save()
        messages.success(request, f'Orden OC-{order.id} cancelada exitosamente.')
        return redirect('suppliers:purchase_order_detail', pk=pk)
    
    return render(request, 'suppliers/purchase_order_confirm_cancel.html', {
        'order': order
    })


@login_required
@role_required(['gerente', 'administrador'])
def purchase_order_export(request):
    """Exportar órdenes de compra a Excel"""
    # Obtener filtros
    supplier_id = request.GET.get('supplier', '')
    status = request.GET.get('status', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    
    # Aplicar filtros
    orders = PurchaseOrder.objects.select_related('supplier').order_by('-created_at')
    
    if supplier_id:
        orders = orders.filter(supplier_id=supplier_id)
    if status:
        orders = orders.filter(status=status)
    if date_from:
        orders = orders.filter(created_at__gte=date_from)
    if date_to:
        orders = orders.filter(created_at__lte=date_to)
    
    # Crear workbook
    wb = openpyxl.Workbook()
    
    # Hoja 1: Resumen de órdenes
    ws1 = wb.active
    ws1.title = "Órdenes de Compra"
    
    # Encabezados
    headers1 = ['N° OC', 'Proveedor', 'Fecha', 'Estado', 'N° Productos', 'Total S/.']
    ws1.append(headers1)
    
    # Estilo para encabezados
    header_fill = PatternFill(start_color="41431B", end_color="41431B", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    for cell in ws1[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Datos de órdenes
    for order in orders:
        items_count = order.items.count()
        status_display = order.get_status_display()
        
        row = [
            f'OC-{order.id}',
            order.supplier.company_name,
            order.created_at.strftime('%d/%m/%Y'),
            status_display,
            items_count,
            float(order.total)
        ]
        ws1.append(row)
    
    # Ajustar ancho de columnas
    ws1.column_dimensions['A'].width = 12
    ws1.column_dimensions['B'].width = 30
    ws1.column_dimensions['C'].width = 12
    ws1.column_dimensions['D'].width = 12
    ws1.column_dimensions['E'].width = 15
    ws1.column_dimensions['F'].width = 15
    
    # Hoja 2: Detalle de items
    ws2 = wb.create_sheet(title="Detalle de Items")
    
    headers2 = ['N° OC', 'Proveedor', 'Producto', 'SKU', 'Cantidad', 'Costo Unit.', 'Subtotal']
    ws2.append(headers2)
    
    for cell in ws2[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Datos de items
    for order in orders:
        for item in order.items.select_related('product').all():
            row = [
                f'OC-{order.id}',
                order.supplier.company_name,
                item.product.name,
                item.product.sku,
                item.quantity,
                float(item.unit_cost),
                float(item.subtotal)
            ]
            ws2.append(row)
    
    # Ajustar ancho de columnas
    ws2.column_dimensions['A'].width = 12
    ws2.column_dimensions['B'].width = 30
    ws2.column_dimensions['C'].width = 30
    ws2.column_dimensions['D'].width = 15
    ws2.column_dimensions['E'].width = 12
    ws2.column_dimensions['F'].width = 15
    ws2.column_dimensions['G'].width = 15
    
    # Preparar respuesta
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'ordenes_compra_ross_crafts_{datetime.now().strftime("%Y%m%d")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename={filename}'
    
    wb.save(response)
    return response
