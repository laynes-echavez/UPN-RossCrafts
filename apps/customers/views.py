from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.views.generic import ListView, CreateView, UpdateView, DetailView
from django.urls import reverse_lazy
from django.http import JsonResponse, HttpResponse
from django.db.models import Q, Sum, Count
from apps.authentication.mixins import RoleRequiredMixin
from apps.authentication.decorators import role_required
from .models import Customer
from .forms import CustomerForm
from apps.sales.models import Sale
from apps.ecommerce.models import Order
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime


class CustomerListView(RoleRequiredMixin, ListView):
    """Lista de clientes con filtros y paginación"""
    allowed_roles = ['gerente', 'administrador', 'empleado']
    model = Customer
    template_name = 'customers/customer_list.html'
    context_object_name = 'customers'
    paginate_by = 25
    
    def get_queryset(self):
        queryset = Customer.objects.order_by('last_name', 'first_name')
        
        # Filtro por búsqueda
        search = self.request.GET.get('search', '')
        if search:
            queryset = queryset.filter(
                Q(first_name__icontains=search) |
                Q(last_name__icontains=search) |
                Q(email__icontains=search) |
                Q(dni__icontains=search)
            )
        
        # Filtro por estado
        is_active = self.request.GET.get('is_active', '')
        if is_active == '1':
            queryset = queryset.filter(is_active=True)
        elif is_active == '0':
            queryset = queryset.filter(is_active=False)
        
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['search'] = self.request.GET.get('search', '')
        context['selected_is_active'] = self.request.GET.get('is_active', '')
        return context


class CustomerCreateView(RoleRequiredMixin, CreateView):
    """Crear nuevo cliente"""
    allowed_roles = ['gerente', 'administrador', 'empleado']
    model = Customer
    form_class = CustomerForm
    template_name = 'customers/customer_form.html'
    success_url = reverse_lazy('customers:customer_list')
    
    def form_valid(self, form):
        messages.success(self.request, f'Cliente "{form.instance.full_name}" registrado exitosamente.')
        return super().form_valid(form)


class CustomerUpdateView(RoleRequiredMixin, UpdateView):
    """Editar cliente existente"""
    allowed_roles = ['gerente', 'administrador', 'empleado']
    model = Customer
    form_class = CustomerForm
    template_name = 'customers/customer_form.html'
    success_url = reverse_lazy('customers:customer_list')
    
    def form_valid(self, form):
        messages.success(self.request, f'Cliente "{form.instance.full_name}" actualizado exitosamente.')
        return super().form_valid(form)


class CustomerProfileView(RoleRequiredMixin, DetailView):
    """Perfil detallado del cliente"""
    allowed_roles = ['gerente', 'administrador', 'empleado']
    model = Customer
    template_name = 'customers/customer_profile.html'
    context_object_name = 'customer'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        customer = self.object
        
        # Historial de ventas presenciales
        sales = Sale.objects.filter(
            customer=customer
        ).select_related('user').order_by('-created_at')
        
        # Historial de pedidos online
        orders = Order.objects.filter(
            customer=customer
        ).order_by('-created_at')
        
        # Estadísticas
        total_sales = sales.aggregate(total=Sum('total'))['total'] or 0
        total_orders = orders.aggregate(total=Sum('total'))['total'] or 0
        total_spent = total_sales + total_orders
        
        total_transactions = sales.count() + orders.count()
        
        # Última compra
        last_sale = sales.first()
        last_order = orders.first()
        
        if last_sale and last_order:
            last_purchase = last_sale if last_sale.created_at > last_order.created_at else last_order
        elif last_sale:
            last_purchase = last_sale
        elif last_order:
            last_purchase = last_order
        else:
            last_purchase = None
        
        context['sales'] = sales[:20]  # Últimas 20 ventas
        context['orders'] = orders[:20]  # Últimos 20 pedidos
        context['total_spent'] = total_spent
        context['total_transactions'] = total_transactions
        context['last_purchase'] = last_purchase
        
        return context


@login_required
@role_required(['gerente', 'administrador'])
def customer_deactivate(request, pk):
    """Desactivar cliente (soft delete)"""
    customer = get_object_or_404(Customer, pk=pk)
    
    if request.method == 'POST':
        customer.is_active = False
        customer.save()
        messages.success(request, f'Cliente "{customer.full_name}" desactivado exitosamente.')
        return redirect('customers:customer_list')
    
    return render(request, 'customers/customer_confirm_deactivate.html', {
        'customer': customer
    })


@login_required
def customer_search_ajax(request):
    """Búsqueda AJAX de clientes para POS"""
    query = request.GET.get('q', '')
    
    if len(query) < 2:
        return JsonResponse({'results': []})
    
    customers = Customer.objects.filter(
        Q(first_name__icontains=query) |
        Q(last_name__icontains=query) |
        Q(email__icontains=query) |
        Q(dni__icontains=query),
        is_active=True
    ).order_by('last_name', 'first_name')[:10]
    
    results = [{
        'id': c.id,
        'full_name': c.full_name,
        'email': c.email,
        'dni': c.dni,
        'phone': c.phone
    } for c in customers]
    
    return JsonResponse({'results': results})


@login_required
@role_required(['gerente', 'administrador'])
def customer_export_excel(request):
    """Exportar clientes a Excel"""
    # Crear workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Clientes"
    
    # Encabezados
    headers = [
        'N°', 'Nombre Completo', 'Email', 'DNI', 'Teléfono',
        'Dirección', 'Total Comprado (S/.)', 'N° Compras',
        'Última Compra', 'Estado'
    ]
    ws.append(headers)
    
    # Estilo para encabezados
    header_fill = PatternFill(start_color="41431B", end_color="41431B", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Obtener clientes con estadísticas
    customers = Customer.objects.order_by('last_name', 'first_name')
    
    for idx, customer in enumerate(customers, start=1):
        # Calcular estadísticas
        sales = Sale.objects.filter(customer=customer)
        orders = Order.objects.filter(customer=customer)
        
        total_sales = sales.aggregate(total=Sum('total'))['total'] or 0
        total_orders = orders.aggregate(total=Sum('total'))['total'] or 0
        total_spent = total_sales + total_orders
        
        total_transactions = sales.count() + orders.count()
        
        # Última compra
        last_sale = sales.order_by('-created_at').first()
        last_order = orders.order_by('-created_at').first()
        
        if last_sale and last_order:
            last_purchase_date = max(last_sale.created_at, last_order.created_at)
        elif last_sale:
            last_purchase_date = last_sale.created_at
        elif last_order:
            last_purchase_date = last_order.created_at
        else:
            last_purchase_date = None
        
        # Agregar fila
        row = [
            idx,
            customer.full_name,
            customer.email,
            customer.dni,
            customer.phone,
            customer.address,
            float(total_spent),
            total_transactions,
            last_purchase_date.strftime('%d/%m/%Y') if last_purchase_date else '-',
            'Activo' if customer.is_active else 'Inactivo'
        ]
        ws.append(row)
    
    # Ajustar ancho de columnas
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 40
    ws.column_dimensions['G'].width = 18
    ws.column_dimensions['H'].width = 12
    ws.column_dimensions['I'].width = 15
    ws.column_dimensions['J'].width = 10
    
    # Preparar respuesta
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'clientes_ross_crafts_{datetime.now().strftime("%Y%m%d")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename={filename}'
    
    wb.save(response)
    return response
