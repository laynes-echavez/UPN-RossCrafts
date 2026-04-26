"""
Vistas para el módulo de auditoría.
Solo accesible por usuarios con rol de gerente.
"""
from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.db.models import Q
from django.http import HttpResponse
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from apps.authentication.decorators import role_required
from apps.authentication.models import User
from .models import AuditLog


@login_required
@role_required(['gerente'])
def audit_log_view(request):
    """Vista principal de auditoría - Solo para gerentes"""
    
    # Obtener parámetros de filtro
    user_filter = request.GET.get('user', '')
    method_filter = request.GET.get('method', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    
    # Query base
    logs = AuditLog.objects.select_related('user').all()
    
    # Aplicar filtros
    if user_filter:
        logs = logs.filter(user_id=user_filter)
    
    if method_filter:
        logs = logs.filter(action=method_filter)
    
    if date_from:
        try:
            date_from_obj = datetime.strptime(date_from, '%Y-%m-%d')
            logs = logs.filter(timestamp__gte=date_from_obj)
        except ValueError:
            pass
    
    if date_to:
        try:
            date_to_obj = datetime.strptime(date_to, '%Y-%m-%d')
            # Incluir todo el día
            from datetime import timedelta
            date_to_obj = date_to_obj + timedelta(days=1)
            logs = logs.filter(timestamp__lt=date_to_obj)
        except ValueError:
            pass
    
    # Ordenar explícitamente para SQL Server
    logs = logs.order_by('-timestamp')
    
    # Limitar a las últimas 500 entradas
    logs = logs[:500]
    
    # Paginación
    paginator = Paginator(logs, 25)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)
    
    # Obtener lista de usuarios para el filtro
    users = User.objects.filter(is_staff=True).order_by('username')
    
    context = {
        'page_obj': page_obj,
        'users': users,
        'user_filter': user_filter,
        'method_filter': method_filter,
        'date_from': date_from,
        'date_to': date_to,
        'total_logs': logs.count(),
    }
    
    return render(request, 'audit/audit_log.html', context)


@login_required
@role_required(['gerente'])
def export_audit_log(request):
    """Exportar registros de auditoría a Excel"""
    
    # Obtener los mismos filtros que la vista principal
    user_filter = request.GET.get('user', '')
    method_filter = request.GET.get('method', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    
    # Query base
    logs = AuditLog.objects.select_related('user').all()
    
    # Aplicar filtros
    if user_filter:
        logs = logs.filter(user_id=user_filter)
    
    if method_filter:
        logs = logs.filter(action=method_filter)
    
    if date_from:
        try:
            date_from_obj = datetime.strptime(date_from, '%Y-%m-%d')
            logs = logs.filter(timestamp__gte=date_from_obj)
        except ValueError:
            pass
    
    if date_to:
        try:
            date_to_obj = datetime.strptime(date_to, '%Y-%m-%d')
            from datetime import timedelta
            date_to_obj = date_to_obj + timedelta(days=1)
            logs = logs.filter(timestamp__lt=date_to_obj)
        except ValueError:
            pass
    
    # Ordenar y limitar
    logs = logs.order_by('-timestamp')[:500]
    
    # Crear libro de Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Auditoría"
    
    # Estilos
    header_fill = PatternFill(start_color="41431B", end_color="41431B", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    # Encabezados
    headers = ['Fecha/Hora', 'Usuario', 'Rol', 'Método', 'URL', 'IP', 'Código HTTP']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Datos
    for row_idx, log in enumerate(logs, start=2):
        ws.cell(row=row_idx, column=1, value=log.timestamp.strftime('%Y-%m-%d %H:%M:%S'))
        ws.cell(row=row_idx, column=2, value=log.user.username if log.user else 'Anónimo')
        ws.cell(row=row_idx, column=3, value=log.user.get_role_display() if log.user else '-')
        ws.cell(row=row_idx, column=4, value=log.action)
        ws.cell(row=row_idx, column=5, value=log.url)
        ws.cell(row=row_idx, column=6, value=log.ip_address)
        ws.cell(row=row_idx, column=7, value=log.status_code)
    
    # Ajustar anchos de columna
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 50
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 12
    
    # Preparar respuesta
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'auditoria_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response
