"""
Script para crear un archivo Excel de ejemplo para importación
"""
import openpyxl
from openpyxl.styles import Font, PatternFill

# Crear workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Productos"

# Encabezados
headers = ['Nombre', 'SKU', 'Categoría', 'Precio', 'Costo', 'Stock Actual', 'Stock Mínimo']
ws.append(headers)

# Estilo para encabezados
header_fill = PatternFill(start_color="AEB784", end_color="AEB784", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF")

for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font

# Datos de ejemplo
productos = [
    ['Bolso Artesanal Grande', 'BOLSO001', 'Textiles', 150.00, 75.00, 20, 5],
    ['Collar de Plata 925', 'COLLAR001', 'Joyería', 200.00, 100.00, 15, 5],
    ['Jarrón de Cerámica Decorativo', 'JARRON001', 'Cerámica', 120.00, 60.00, 10, 3],
    ['Tapiz Decorativo Andino', 'TAPIZ001', 'Decoración', 180.00, 90.00, 8, 3],
    ['Pulsera Artesanal Tejida', 'PULSERA001', 'Joyería', 80.00, 40.00, 30, 10],
    ['Plato Decorativo Pintado', 'PLATO001', 'Cerámica', 60.00, 30.00, 25, 8],
    ['Bufanda Tejida a Mano', 'BUFANDA001', 'Textiles', 90.00, 45.00, 18, 5],
    ['Figura Decorativa Madera', 'FIGURA001', 'Artesanías', 110.00, 55.00, 12, 4],
    ['Aretes de Plata', 'ARETES001', 'Joyería', 70.00, 35.00, 40, 15],
    ['Mantel Bordado', 'MANTEL001', 'Textiles', 130.00, 65.00, 10, 3],
]

for producto in productos:
    ws.append(producto)

# Ajustar ancho de columnas
ws.column_dimensions['A'].width = 30
ws.column_dimensions['B'].width = 15
ws.column_dimensions['C'].width = 15
ws.column_dimensions['D'].width = 12
ws.column_dimensions['E'].width = 12
ws.column_dimensions['F'].width = 15
ws.column_dimensions['G'].width = 15

# Guardar archivo
wb.save('ejemplo_importacion_productos.xlsx')

print("✓ Archivo 'ejemplo_importacion_productos.xlsx' creado exitosamente")
print(f"  {len(productos)} productos de ejemplo incluidos")
print("\nPuedes usar este archivo para probar la importación en:")
print("  http://localhost:8000/stock/importar/")
